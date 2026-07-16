import json
import uuid
from urllib.parse import parse_qs, urlparse
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.template.loader import get_template
from django.conf import settings
from django.utils import timezone
from django.db.models import Count
from invitations.models import Invitation
from .models import Guest
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from django.db import transaction
from .whatsapp_messages import (
    get_whatsapp_message_options,
    get_whatsapp_message_template,
    render_whatsapp_message,
)

@require_POST
def submit_rsvp(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)
    
    # Check if there is a guest token
    token = request.POST.get('guest_token')
    
    effective_deadline = invitation.rsvp_deadline
    guest_obj = None
    if token:
        try:
            guest_obj = Guest.objects.get(token=token, invitation=invitation)
            if guest_obj.rsvp_deadline:
                effective_deadline = guest_obj.rsvp_deadline
        except Guest.DoesNotExist:
            pass
            
    if effective_deadline and timezone.now() > effective_deadline:
        return JsonResponse({'status': 'error', 'message': 'La fecha límite de confirmación ha expirado.'}, status=400)
    
    guest_name = request.POST.get('guest_name')
    attending = request.POST.get('attending') == 'true'
    
    try:
        number_of_companions = int(request.POST.get('number_of_companions', 0))
    except ValueError:
        number_of_companions = 0
        
    comments = request.POST.get('comments', '')
    
    if token:
        # It's an invited guest
        if not guest_obj:
            return JsonResponse({'status': 'error', 'message': 'Token de invitado no válido.'}, status=400)
            
        if attending and number_of_companions > guest_obj.max_companions:
            return JsonResponse({
                'status': 'error', 
                'message': f'Solo puedes confirmar hasta {guest_obj.max_companions} acompañantes.'
            }, status=400)
            
        guest_obj.has_responded = True
        guest_obj.is_attending = attending
        guest_obj.confirmed_companions = number_of_companions if attending else 0
        guest_obj.dietary_restrictions = f"Comentarios: {comments}" if comments else ""
        guest_obj.save()
        
    else:
        return JsonResponse({
            'status': 'error',
            'message': 'Necesitas un enlace personalizado para confirmar asistencia.'
        }, status=400)
        
    return JsonResponse({
        'status': 'success',
        'message': '¡Gracias por confirmar tu asistencia!',
        'guest': {
            'name': guest_obj.name,
            'alias': guest_obj.alias or '',
            'has_responded': guest_obj.has_responded,
            'is_attending': guest_obj.is_attending,
            'confirmed_companions': guest_obj.confirmed_companions,
            'max_companions': guest_obj.max_companions,
        },
    })


@staff_member_required
def generate_guest_pdf(request, guest_id):
    guest = get_object_or_404(Guest, id=guest_id)
    invitation = guest.invitation
    
    # Message logic
    if invitation.default_pdf_message:
        message = invitation.default_pdf_message
    else:
        message = "Te invitamos cordialmente a celebrar este día tan especial con nosotros. Tu presencia es el mejor regalo."
        
    # Get the URL for the pass
    protocol = "https" if request.is_secure() else "http"
    domain = request.get_host()
    link = f"{protocol}://{domain}/{invitation.slug}/?guest={guest.token}"
    
    if invitation.section_background_image:
        background_url = f"{protocol}://{domain}{settings.MEDIA_URL}{invitation.section_background_image.name}"
    else:
        background_url = f"{protocol}://{domain}{settings.STATIC_URL}img/fondo_floral.png"
    
    context = {
        'guest': guest,
        'invitation': invitation,
        'message': message,
        'link': link,
        'background_url': background_url,
        'MEDIA_URL': f"{protocol}://{domain}{settings.MEDIA_URL}" if settings.MEDIA_URL else "",
        'STATIC_URL': f"{protocol}://{domain}{settings.STATIC_URL}" if settings.STATIC_URL else ""
    }
    
    template_path = 'rsvp/guest_pdf.html'
    template = get_template(template_path)
    html = template.render(context)
    
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f"Pase_{guest.name.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    
    return HttpResponse('Error generating PDF', status=400)


from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from functools import wraps

ADMIN_GROUP_NAME = 'Administrador de invitaciones'
HOSTESS_GROUP_NAME = 'Hostess / Recepcionista'


def user_is_invitation_admin(user, invitation):
    if not user.is_authenticated:
        return False
    if user.is_staff or user.is_superuser:
        return True
    return invitation.administrators.filter(pk=user.pk).exists()


def user_is_invitation_hostess(user, invitation):
    if not user.is_authenticated:
        return False
    if user_is_invitation_admin(user, invitation):
        return True
    return invitation.hostesses.filter(pk=user.pk).exists()

def check_invitation_access(view_func):
    @login_required
    @wraps(view_func)
    def _wrapped_view(request, slug, *args, **kwargs):
        invitation = get_object_or_404(Invitation, slug=slug)
        if not user_is_invitation_admin(request.user, invitation):
            raise PermissionDenied("No tienes permiso para administrar esta invitación.")
        return view_func(request, slug, *args, **kwargs)
    return _wrapped_view


def check_reception_access(view_func):
    @login_required
    @wraps(view_func)
    def _wrapped_view(request, slug, *args, **kwargs):
        invitation = get_object_or_404(Invitation, slug=slug)
        if not user_is_invitation_hostess(request.user, invitation):
            raise PermissionDenied("No tienes permiso para validar entradas de esta invitación.")
        return view_func(request, slug, *args, **kwargs)
    return _wrapped_view

# --- GUEST DASHBOARD ENDPOINTS ---

def is_mobile_request(request):
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    if not user_agent:
        return False

    tablet_indicators = ('ipad', 'tablet')
    if any(indicator in user_agent for indicator in tablet_indicators):
        return False

    if 'android' in user_agent and 'mobile' not in user_agent:
        return False

    mobile_indicators = (
        'iphone',
        'ipod',
        'android',
        'windows phone',
        'blackberry',
        'opera mini',
        'iemobile',
    )
    return any(indicator in user_agent for indicator in mobile_indicators)


@check_invitation_access
@ensure_csrf_cookie
def guest_dashboard(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)
    template_name = 'rsvp_mobile_guest_dashboard.html' if is_mobile_request(request) else 'rsvp_guest_dashboard.html'
    return render(request, template_name, {'invitation': invitation})


@check_reception_access
@ensure_csrf_cookie
def hostess_dashboard(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)
    return render(request, 'rsvp_hostess_dashboard.html', {
        'invitation': invitation,
        'can_manage_guests': user_is_invitation_admin(request.user, invitation),
    })


@check_invitation_access
@require_http_methods(["GET", "POST"])
def api_guests_list_create(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)
    
    if request.method == "GET":
        guests = Guest.objects.filter(invitation=invitation).annotate(visit_count=Count('visits')).order_by('-created_at')
        data = [{
            'id': g.id,
            'name': g.name,
            'alias': g.alias,
            'phone_number': g.phone_number,
            'max_companions': g.max_companions,
            'has_responded': g.has_responded,
            'is_attending': g.is_attending,
            'confirmed_companions': g.confirmed_companions,
            'token': g.token,
            'sheet_name': g.sheet_name,
            'whatsapp_sent': g.whatsapp_sent,
            'visit_count': g.visit_count,
            'checked_in_at': g.checked_in_at.isoformat() if g.checked_in_at else None,
            'checked_in_by_name': g.checked_in_by.get_username() if g.checked_in_by else '',
            'check_in_method': g.check_in_method,
        } for g in guests]
        return JsonResponse(data, safe=False)
        
    elif request.method == "POST":
        try:
            body = json.loads(request.body)
            guest = Guest.objects.create(
                invitation=invitation,
                name=body.get('name'),
                alias=body.get('alias', ''),
                phone_number=body.get('phone_number', ''),
                max_companions=int(body.get('max_companions', 1))
            )
            return JsonResponse({'status': 'success', 'id': guest.id}, status=201)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@check_invitation_access
@require_http_methods(["PUT", "DELETE"])
def api_guest_detail(request, slug, guest_id):
    invitation = get_object_or_404(Invitation, slug=slug)
    guest = get_object_or_404(Guest, id=guest_id, invitation=invitation)
    
    if request.method == "PUT":
        try:
            body = json.loads(request.body)
            guest.name = body.get('name', guest.name)
            guest.alias = body.get('alias', guest.alias)
            guest.phone_number = body.get('phone_number', guest.phone_number)
            guest.max_companions = int(body.get('max_companions', guest.max_companions))
            guest.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    elif request.method == "DELETE":
        guest.delete()
        return JsonResponse({'status': 'success'}, status=204)


@check_invitation_access
@require_http_methods(["POST"])
def api_guest_whatsapp_sent(request, slug, guest_id):
    invitation = get_object_or_404(Invitation, slug=slug)
    guest = get_object_or_404(Guest, id=guest_id, invitation=invitation)
    if not guest.whatsapp_sent:
        guest.whatsapp_sent = True
        guest.save(update_fields=['whatsapp_sent', 'updated_at'])
    return JsonResponse({'status': 'success', 'whatsapp_sent': guest.whatsapp_sent})


@check_invitation_access
@require_http_methods(["GET"])
def api_whatsapp_messages(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)
    return JsonResponse(get_whatsapp_message_options(invitation), safe=False)


@check_invitation_access
@require_http_methods(["GET"])
def api_guest_whatsapp_message_preview(request, slug, guest_id):
    invitation = get_object_or_404(Invitation, slug=slug)
    guest = get_object_or_404(Guest, id=guest_id, invitation=invitation)
    message_id = request.GET.get('message_id') or None
    template = get_whatsapp_message_template(invitation, message_id)
    return JsonResponse({'message': render_whatsapp_message(template, invitation, guest, request)})


@check_invitation_access
@require_http_methods(["GET"])
def api_download_excel_template(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invitados"
    
    # Headers
    headers = ["Nombre del Invitado o Familia", "Alias", "Teléfono (WhatsApp)", "Número de Pases (Lugares)"]
    ws.append(headers)
    
    # Styling headers
    header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    
    # Add an example row
    ws.append(["Familia Ejemplo", "Tíos de Monterrey", "5215512345678", 4])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Invitados.xlsx"'
    wb.save(response)
    
    return response


@check_invitation_access
@require_http_methods(["POST"])
def api_upload_excel_guests(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)
    
    if 'file' not in request.FILES:
        return JsonResponse({'status': 'error', 'message': 'No se encontró archivo'}, status=400)
        
    excel_file = request.FILES['file']
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': 'Archivo Excel inválido o corrupto.'}, status=400)
        
    # Validation and Atomic Transaction
    created_count = 0
    updated_count = 0
    
    try:
        with transaction.atomic():
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_values = [str(cell.value or '').strip().lower() for cell in ws[1]]
                has_alias_column = len(header_values) > 1 and 'alias' in header_values[1]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue  # Skip empty rows or rows without name
                        
                    name = str(row[0]).strip()
                    alias = ""
                    if has_alias_column and len(row) > 1 and row[1] is not None:
                        alias = str(row[1]).strip()
                    
                    # Parse phone carefully to remove decimal points if stored as numeric float in Excel
                    phone = ""
                    phone_index = 2 if has_alias_column else 1
                    if len(row) > phone_index and row[phone_index] is not None:
                        val = row[phone_index]
                        if isinstance(val, float):
                            phone = str(int(val)) if val.is_integer() else str(val)
                        elif isinstance(val, int):
                            phone = str(val)
                        else:
                            phone = str(val).strip()
                            if phone.endswith('.0'):
                                phone = phone[:-2]
                    
                    # Try to parse passes
                    passes_index = 3 if has_alias_column else 2
                    try:
                        passes = int(row[passes_index])
                    except (IndexError, ValueError, TypeError):
                        passes = 1  # Default to 1 pass if invalid
                        
                    if not name:
                        continue
                        
                    # Search for duplicate: by phone first, then by name
                    guest = None
                    if phone:
                        guest = Guest.objects.filter(invitation=invitation, phone_number=phone).first()
                    if not guest:
                        guest = Guest.objects.filter(invitation=invitation, name=name).first()
                        
                    if guest:
                        # Update existing guest
                        guest.name = name
                        guest.alias = alias
                        if phone:
                            guest.phone_number = phone
                        guest.max_companions = passes
                        guest.sheet_name = sheet_name
                        guest.created_by = request.user
                        guest.save()
                        updated_count += 1
                    else:
                        # Create new guest
                        Guest.objects.create(
                            invitation=invitation,
                            name=name,
                            alias=alias,
                            phone_number=phone,
                            max_companions=passes,
                            sheet_name=sheet_name,
                            created_by=request.user
                        )
                        created_count += 1
                    
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error procesando el archivo: {str(e)}'}, status=400)
        
    return JsonResponse({
        'status': 'success', 
        'message': f'Importación completada: {created_count} creados, {updated_count} actualizados.'
    })


@check_invitation_access
@require_http_methods(["GET", "POST"])
def api_hostesses_list_create(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)
    User = get_user_model()

    if request.method == "GET":
        data = [{
            'id': user.id,
            'username': user.get_username(),
            'first_name': user.first_name,
            'last_name': user.last_name,
        } for user in invitation.hostesses.order_by('username')]
        return JsonResponse(data, safe=False)

    try:
        body = json.loads(request.body)
        username = (body.get('username') or '').strip()
        password = body.get('password') or ''
        first_name = (body.get('first_name') or '').strip()
        last_name = (body.get('last_name') or '').strip()

        if not username:
            return JsonResponse({'status': 'error', 'message': 'El usuario es obligatorio.'}, status=400)

        user = User.objects.filter(username=username).first()
        created = False
        if user is None:
            if not password:
                return JsonResponse({'status': 'error', 'message': 'La contraseña temporal es obligatoria para usuarios nuevos.'}, status=400)
            user = User.objects.create_user(username=username, password=password, first_name=first_name, last_name=last_name)
            created = True
        else:
            if first_name:
                user.first_name = first_name
            if last_name:
                user.last_name = last_name
            if first_name or last_name:
                user.save(update_fields=['first_name', 'last_name'])

        hostess_group, _ = Group.objects.get_or_create(name=HOSTESS_GROUP_NAME)
        user.groups.add(hostess_group)
        invitation.hostesses.add(user)

        return JsonResponse({
            'status': 'success',
            'created': created,
            'hostess': {
                'id': user.id,
                'username': user.get_username(),
                'first_name': user.first_name,
                'last_name': user.last_name,
            },
        }, status=201 if created else 200)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido.'}, status=400)


@check_invitation_access
@require_http_methods(["DELETE"])
def api_hostess_detail(request, slug, user_id):
    invitation = get_object_or_404(Invitation, slug=slug)
    User = get_user_model()
    user = get_object_or_404(User, pk=user_id)
    invitation.hostesses.remove(user)
    return JsonResponse({'status': 'success'}, status=200)


def serialize_reception_guest(guest):
    return {
        'id': guest.id,
        'name': guest.name,
        'alias': guest.alias,
        'phone_number': guest.phone_number,
        'max_companions': guest.max_companions,
        'has_responded': guest.has_responded,
        'is_attending': guest.is_attending,
        'confirmed_companions': guest.confirmed_companions,
        'checked_in_at': guest.checked_in_at.isoformat() if guest.checked_in_at else None,
        'checked_in_by_name': guest.checked_in_by.get_username() if guest.checked_in_by else '',
        'check_in_method': guest.check_in_method,
    }


def extract_guest_token(raw_value):
    value = (raw_value or '').strip()
    if not value:
        return None

    parsed = urlparse(value)
    query = parse_qs(parsed.query)
    if query.get('guest'):
        value = query['guest'][0]

    try:
        return uuid.UUID(str(value))
    except (TypeError, ValueError):
        return None


@check_reception_access
@require_http_methods(["GET"])
def api_hostess_guests(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)
    guests = Guest.objects.filter(
        invitation=invitation,
        has_responded=True,
        is_attending=True,
    ).order_by('name')
    return JsonResponse([serialize_reception_guest(guest) for guest in guests], safe=False)


@check_reception_access
@require_http_methods(["POST"])
def api_hostess_check_in(request, slug):
    invitation = get_object_or_404(Invitation, slug=slug)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido.'}, status=400)

    guest = None
    method = body.get('method') or 'manual'
    notes = body.get('notes') or ''

    guest_id = body.get('guest_id')
    if guest_id:
        guest = Guest.objects.filter(invitation=invitation, pk=guest_id).first()
    else:
        token = extract_guest_token(body.get('token_or_url') or body.get('token'))
        if token:
            guest = Guest.objects.filter(invitation=invitation, token=token).first()

    if guest is None:
        return JsonResponse({'status': 'error', 'message': 'No se encontró un invitado válido para esta invitación.'}, status=404)

    if not guest.has_responded or not guest.is_attending:
        return JsonResponse({
            'status': 'error',
            'message': 'No se permite la entrada porque este invitado no hizo la confirmación correspondiente.',
            'guest': serialize_reception_guest(guest),
        }, status=403)

    if guest.checked_in_at:
        return JsonResponse({
            'status': 'duplicate',
            'message': 'Este invitado ya fue validado previamente.',
            'guest': serialize_reception_guest(guest),
        }, status=200)

    guest.checked_in_at = timezone.now()
    guest.checked_in_by = request.user
    guest.check_in_method = 'qr' if method == 'qr' else 'manual'
    guest.check_in_notes = notes
    guest.save(update_fields=['checked_in_at', 'checked_in_by', 'check_in_method', 'check_in_notes', 'updated_at'])

    return JsonResponse({
        'status': 'success',
        'message': 'Entrada validada correctamente.',
        'guest': serialize_reception_guest(guest),
    })
